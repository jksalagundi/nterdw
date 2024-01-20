from openpyxl import load_workbook
import sys
from datetime import datetime
from orders.bookings.data_upsert import upsert_booking_data
import logging

ARRIVAL_DATE_FORMAT = '%m/%d/%y %I:%M%p'
BOOKING_DATE_FORMAT = '%m/%d/%y, %I:%M %p'
logger = logging.getLogger('nterdw')


class BookingProcessException(Exception):
    pass


def __get_val(sheet, row, col, default=None):
    value = sheet.cell(row=row, column=col).value
    return value if value is not None else default


def get_booking_record(sheet, row, location):
    """
    Given the worksheet and the row, and return the booking record
    """
    try:
        booking_id = sheet.cell(row=row, column=1).value
        listing_name = sheet.cell(row=row, column=2).value
        arrival_date = sheet.cell(row=row, column=3).value
        arrival_time = __get_val(sheet, row, 4, '9:00am')
        arrival_dt = datetime.strptime(f"{arrival_date} {arrival_time}", ARRIVAL_DATE_FORMAT)
        customer_name = sheet.cell(row=row, column=5).value
        customer_email = sheet.cell(row=row, column=6).value
        guests = int(__get_val(sheet, row, 7, 0))
        base_amount = float(__get_val(sheet, row, 8, 0.0))
        taxes = float(__get_val(sheet, row, 9, 0.0))
        addons = int(__get_val(sheet, row, 10, 0))
        coupon_code = sheet.cell(row=row, column=11).value
        amount = float(__get_val(sheet, row, 12, 0.0))
        adjustments = float(__get_val(sheet, row, 13, 0.0))
        total_value = float(__get_val(sheet, row, 14, 0.0))
        revenue = float(__get_val(sheet, row, 15, 0.0))
        # Skipping Balance Due column 16
        source = sheet.cell(row=row, column=17).value
        status = sheet.cell(row=row, column=18).value
        # Skipping Guides & Guide Status columns
        booking_date = datetime.strptime(sheet.cell(row=row, column=21).value, BOOKING_DATE_FORMAT)
        reserved_by = sheet.cell(row=row, column=22).value

        return {
            'booking_id': booking_id,
            'listing_name': listing_name,
            'arrival_date_time': arrival_dt,
            'customer_name': customer_name,
            'customer_email': customer_email,
            'guest_count': guests,
            'base_amount': base_amount,
            'taxes': taxes,
            'addons': addons,
            'coupon_code': coupon_code,
            'amount': amount,
            'adjustments': adjustments,
            'total_value': total_value,
            'revenue': revenue,
            'source': source,
            'status': status,
            'booking_date': booking_date,
            'reserved_by': reserved_by,
            'location': location
        }
    except ValueError as vex:
        logger.warning(f"Exception while converting data from string to another format {vex.__str__()}. "
                       f"BookingID: {sheet.cell(row=row, column=1).value}")
        raise BookingProcessException(f"Exception while converting string to integer {vex}")
    except Exception as ex:
        logger.warning(f"General Exception while converting data from string to another format {ex.__str__()}. "
                       f"BookingID: {sheet.cell(row=row, column=1).value}")
        raise BookingProcessException(ex)


def load_bookings_data(bookings_file, location):

    if bookings_file is None:
        print('Invalid information - No bookings file found to process')
        sys.exit(-1)

    workbook = load_workbook(filename=bookings_file)
    logger.debug(f"Worksheet names: {workbook.sheetnames}")

    sheet = workbook.get_sheet_by_name('Bookings')
    logger.debug(f"Workbook title to process {sheet.title} with max rows {sheet.max_row}")

    for row in range(3, sheet.max_row):
        booking_rec = get_booking_record(sheet, row, location)
        upsert_booking_data(booking_rec)
        # if row > 5:
        #     break



